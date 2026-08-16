from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_table_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"


def _fit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width, 40)


def _write_dataframe(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.append(dataframe.columns.tolist())
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append(list(row))
    _style_table_header(worksheet)
    _fit_columns(worksheet)


def build_excel_report(
    df: pd.DataFrame,
    metrics: dict[str, float | int],
    performance: pd.DataFrame,
    insights: list[str],
) -> bytes:
    """Create a downloadable multi-sheet Excel business report."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Executive Summary"

    summary["A1"] = "InsightIQ Business Report"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = HEADER_FILL
    summary.merge_cells("A1:B1")
    summary["A3"] = "Report period"
    summary["B3"] = f"{df['Date'].min().date()} to {df['Date'].max().date()}"
    summary["A5"] = "Total Revenue"
    summary["B5"] = metrics["revenue"]
    summary["A6"] = "Orders"
    summary["B6"] = metrics["orders"]
    summary["A7"] = "Items Sold"
    summary["B7"] = metrics["items"]
    summary["A8"] = "Average Order"
    summary["B8"] = metrics["average_order"]
    summary["B5"].number_format = "$#,##0.00"
    summary["B8"].number_format = "$#,##0.00"
    summary["A10"] = "AI Insights"
    summary["A10"].font = Font(bold=True)
    for index, insight in enumerate(insights, start=11):
        summary[f"A{index}"] = insight
        summary.merge_cells(start_row=index, start_column=1, end_row=index, end_column=2)
        summary[f"A{index}"].alignment = Alignment(wrap_text=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 50

    performance_sheet = workbook.create_sheet("Product Performance")
    _write_dataframe(performance_sheet, performance)
    for row in range(2, performance_sheet.max_row + 1):
        performance_sheet[f"B{row}"].number_format = "$#,##0.00"

    sales_sheet = workbook.create_sheet("Filtered Sales Data")
    export_df = df.copy()
    export_df["Date"] = export_df["Date"].dt.strftime("%Y-%m-%d")
    _write_dataframe(sales_sheet, export_df)
    revenue_column = export_df.columns.get_loc("Revenue") + 1
    for row in range(2, sales_sheet.max_row + 1):
        sales_sheet.cell(row=row, column=revenue_column).number_format = "$#,##0.00"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
